# coding: utf-8
team_list = ['O&M 1','O&M 2', 'BJ3 O&M2','O&M 5','OAM Dev Sigma&Puma','OAM Dev Y&E','OAM Dev Z1&Z2','OAM Dev Wave','BJ3 O&M1', 'T&P']
status_list = ['Community Discussion', 'Implementation started', 'Implementation done']
date_range = ['20170101', '20171231']

row_labels=status_list
row_labels.append("Innovation Point")
#row_labels = ["Discussion", "Implemented Started", "Implementation Done", "Total Score"] #for output and draw

import pandas as pd
import re
import matplotlib.pyplot as plt
import datetime as dt
import numpy as np
from matplotlib.ticker import MaxNLocator
from collections import namedtuple
import json
from openpyxl import (
    Workbook,
    load_workbook
)
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
    BarChart
)

#load excel file
#files=[
#    r'D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\20171224-export\\p0.xlsx',
#    r'D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\20171224-export\\p1.xlsx',
#    r'D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\20171224-export\\p1.xlsx'
#]

files=[
    r'D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\2017_export\\p_all.xlsx',
    ]

compactCsv='D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\2017_export\\compactView.xlsx'
statisticCsv='D:\\LTE\\Nokia\\TC innovation\\0-MONO7\\2017_export\\statisticView.xlsx'

#initiate global variables
idea_numbers=[]
idea_titles=[]
idea_initialtors=[]
idea_funnels=[]
idea_status=[]
community_discussion=[]
implementation_started=[]
implementation_done=[]
datetime_list=[]

'''
SCORES = [
    [team1, team2, team3, ...], #Community Discussion 
    [team1, team2, team3, ...], #Implementation started 
    [team1, team2, team3, ...]  #Implementation done
    ]

'''
SCORES=np.zeros((4, len(team_list)), dtype=np.int16)

def addData(file):
    '''
        read file and store the data
    '''    
    data = pd.read_excel(file)

    #get global variable
    global idea_numbers, idea_titles, idea_initialtors, idea_funnels, idea_status, community_discussion, implementation_started, implementation_done
    idea_numbers=data['Idea number']
    idea_titles=data['Idea Title']
    idea_initialtors=data['Initiator']
    idea_funnels=data['Idea Funnel']
    idea_status=data['Idea Status']
    community_discussion=data['Community Discussion']
    implementation_started=data['Implementation']
    implementation_done=data['Implementation done']
    print("Read data from file:", file, ", Number of ideas:", len(idea_numbers), "\n")
    
#construct datetime from date_range
def construct_datetimes_range():
    for i in range(len(date_range)):
        year = int(date_range[i][0:4])
        month = int(date_range[i][4:6])
        day = int(date_range[i][6:8])
        datetime_list.append(dt.datetime(year, month, day, 0, 0, 0))

def drawTable(scores):
    data = scores.tolist()
    x_groups = np.amax(scores)
    yvalues = np.arange(4)
    xvalues = np.arange(len(data))

    #data = [[ 66386, 174296,  75131, 577908,  32015],
    #   [ 58230, 381139,  78045,  99308, 160454],
    #   [ 89135,  80552, 152558, 497981, 603535],
    #   [ 78415,  81858, 150656, 193263,  69638],
    #   [139361, 331509, 343164, 781380,  52269]]

    columns = team_list
    rows = ['%d' % x for x in yvalues]
    
    print("rows:", rows)
    #row_lables = ["Discussion", "Implemented Started", "Implementation Done", "Total Score"]

    values = np.arange(0,  x_groups, 1)
    value_increment = 1

    # Get some pastel shades for the colors
    colors = plt.cm.BuPu(np.linspace(0, 0.5, len(rows)))
    n_rows = len(data)

    index = np.arange(len(columns)) + 0.3
    bar_width = 0.4

    # Initialize the vertical-offset for the stacked bar chart.
    y_offset = np.zeros(len(columns))

    # Plot bars and create text labels for the table
    cell_text = []
    for row in range(n_rows):
        plt.bar(index, data[row], bar_width, bottom=y_offset, color=colors[row])
        y_offset = y_offset + data[row]
        cell_text.append(['%1.1f' % (x) for x in y_offset])
    # Reverse colors and text labels to display the last value at the top.
    colors = colors[::-1]
    cell_text.reverse()

    # Add a table at the bottom of the axes
    the_table = plt.table(cellText=cell_text,
                          rowLabels=row_labels,
                          rowColours=colors,
                          colLabels=columns,
                          loc='bottom')

    # Adjust layout to make room for the table:
    plt.subplots_adjust(left=0.2, bottom=0.2)

    plt.ylabel("Loss in ${0}'s".format(value_increment))
    plt.yticks(values * value_increment, ['%d' % val for val in values])
    plt.xticks([])
    plt.title('Loss by Disaster')

    plt.show()

def draw(scores):
    #draw the bar
    n_groups = len(team_list)
    x_groups = np.amax(scores)
    print(x_groups)

    discussion = scores[0]
    implements = scores[1]
    done = scores[2]

    fig, ax= plt.subplots()

    index = np.arange(n_groups )
    print ("index:", index)

    yvalues = np.arange(x_groups)
    print("yvalues:", yvalues)
    yvalue_increment = 1
    
    bar_width = 0.25

    opacity = 0.8

    rects1 = ax.bar(index-bar_width/2, discussion, bar_width,
                    alpha=opacity, color='b',
                    label='Community Discussion')

    rects2 = ax.bar(index + bar_width/2, implements, bar_width,
                    alpha=opacity, color='r',
                    label='Implementaed Started')

    rects3 = ax.bar(index + bar_width*1.5, done, bar_width,
                    alpha=opacity, color='y',
                    label='Implementation Done', log=True)

    ax.set_xlabel('Tribe')
    ax.set_ylabel('Scores')
    ax.set_title('Scores by Tribe ')
    ax.set_xticks(index + bar_width*2)    

    #print(score_range)
    ax.set_xticklabels(team_list)

    #print(ind)
    #ax.set_yticks(ind)
    ax.set_yticklabels(yvalues)

    ax.legend()

    plt.setp(ax.xaxis.get_majorticklabels(), rotation=-50 )

    fig.tight_layout()
    
    the_table = plt.table(cellText=scores,
                          rowLabels=row_labels,
                          #rowColours=colors,
                          colLabels=team_list,
                          loc='bottom')

    # Adjust layout to make room for the table:
    plt.subplots_adjust(left=0.2, bottom=0.2)
    
    plt.show()

def format_excel(writer):
    """ Add Excel specific formatting to the workbook
    """
    # Get the workbook and the summary sheet so we can add the formatting
    workbook = writer.book
    worksheet = writer.sheets['summary']
    # Add currency formatting and apply it
    money_fmt = workbook.add_format({'num_format': 42, 'align': 'center'})
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:C', 15, money_fmt)
    worksheet.add_table('A1:C22', {'columns': [{'header': 'account',
                                                'total_string': 'Total'},
                                               {'header': 'Total Sales',
                                                'total_function': 'sum'},
                                               {'header': 'Average Sales',
                                                'total_function': 'average'}],
                                   'autofilter': False,
                                   'total_row': True,
                                   'style': 'Table Style Medium 20'})
                                   
def saveCompactCsv(file, ideas):    
    df = pd.DataFrame((idea.toJson() for idea in ideas))
    
    #using workbook openpyxl ===> Yan, it looks better than pandas, to study it
    from openpyxl import Workbook
    wb = Workbook()
    wb.create_sheet("CampactView") 
    #wb.save(file)
    
    #print("OOOO:", df)
    #df.to_excel(file, "CampactView") 
    writer = pd.ExcelWriter(file, engine='xlsxwriter')
    df.to_excel(writer, 'CampactView', index=True)
    #format_excel(writer)
    writer.save()    
    print("Compact view of all ideas is stored in: ", file)
        
def saveStatisticCsv(file, scores, ideas):

    # {'A-Community Discussion': score, 'B-Implementation Started' : score[1],'C-Idea Status': self.idea_status,'D-Idea Date:':self.idea_date,  'B-Idea Funnel':self.idea_funnel }
    scorelist=scores.tolist()
    #columns=team_list
    #index=row_labels 
    #dict={}
    
    #dict={'A': '1', 'B': '2'}
    #for i in range(len(row_labels)):
    #    for j, label in enumerate(team_list):
    #        dict[i][label] = scorelist[i`][j]
        
    #jsonScores=json.dumps(dict)
    jsonScores = {'A':['1'], 'B': ['2']}
    #print(jsonScores)
    
    #using workbook openpyxl ===> Yan, it looks better than pandas, to study it
    wb = Workbook(write_only=True)
    ws_sta=wb.create_sheet("StatisticView") 
    wb.create_sheet("CampactView") 
    
    df = pd.DataFrame(scorelist, columns=team_list, index=row_labels) #, sheet_name="StatisticView")
    writer = pd.ExcelWriter(file, engine='xlsxwriter')
    df.to_excel(writer, 'StatisticView', index=True)
    
    ws_sta.append(team_list)
    for score in scores:
        ws_sta.append(score.tolist()) 
    
    #write ideas view:
    df1 = pd.DataFrame((idea.toJson() for idea in ideas))
    df1.to_excel(writer, 'CampactView', index=True)
    wb.save(file)
        
    #writer.save()    
    #wb=load_workbook(file)
    print("Statistic view of all ideas is stored in: ", file)
 
def testChart(file, scores): 
    #have a test
    wb = Workbook(write_only=True)
    ws_sta = wb.create_sheet()

    #header = [
    #    ' ', 'Discussion', 'Implementation started', 'Implementation Done', 'Total Innovation Point'
    #]
    
    headers=row_labels;
    headers.insert(0, ' ')

    revertrows=list()
    for i in range(len(team_list)):
        revertrow = [x[i] for x in scores]
        revertrow.insert(0, team_list[i])
        print("revertrow:", revertrow)
        revertrows.append(revertrow)
    
    print("revertrows:", revertrows)
    
    ws_sta.append(headers)
    for row in revertrows:
        print(row)
        ws_sta.append(row)

    data = Reference(ws_sta, min_col=2, min_row=1, max_col=4, max_row=len(team_list)+1)
    titles = Reference(ws_sta, min_col=1, min_row=2, max_row=len(team_list)+1)
    print("data:", data)
    print("title:", titles)
    chart = BarChart3D()
    chart.title = "Innovation Statistic - Idea status"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    ws_sta.add_chart(chart, "A15")
    
    data = Reference(ws_sta, min_col=5, min_row=1, max_col=5, max_row=len(team_list)+1)
    titles = Reference(ws_sta, min_col=1, min_row=2, max_row=len(team_list)+1)
    print("data:", data)
    print("title:", titles)
    chart = BarChart3D()
    chart.title = "Innovation Statistic - Innovation Point"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    ws_sta.add_chart(chart, "I15")
    
    wb.save(file)
    
    
class CIdea:
    global team_list, datetime_list, status_list
    def __init__(self, number, title, initiator, funnel, status, date_discussion, date_started, date_done ):
        self.idea_number = number
        self.idea_title = title
        self.idea_initialtor = initiator
        self.idea_funnel = self.getShortName(funnel)
        self.idea_status = status
        self.idea_date = dt.datetime(1901, 1, 1, 0, 0, 0)
        
        if status == 'Community Discussion':
            #print("this is in : ", status, " on ", date_discussion)
            self.idea_date = date_discussion
        if status == "Implementation started":
            #print("This idea is in :", status, " on ", date_started)
            self.idea_date = date_started
        if status == "Implementation done":
            #print("This idea is in :", status, " on ", date_done)
            self.idea_date = date_done
    
    def __eq__(self, another):
        return self.idea_number == another.idea_number
    
    def __hash__(self):
        return hash(self.idea_number)
        
    def getNumber(self):
        return self.idea_number
        
    def getFunnel(self):
        return idea_funnel
    
    def getStatus(self):
        return idea_status
    
    def getDate(self):
        return idea_date
    
    #get short funnel name
    def getShortName(self, longName):
        for i in range(len(team_list)):
            shortName=team_list[i]
            #print("short name:", shortName, ",longName:", longName)
            if shortName.strip() in longName:
                #print("Use shortName:", shortName)
                return shortName
        
        print("Use longName:", longName)
        return longName

    def isInDateRange(self):
        if self.idea_date < datetime_list[1] and self.idea_date > datetime_list[0]:
            #print ("this idea is to be counted: ", self.idea_date)
            return True
        else:
            return False
    
    def isValid(self):
        isValid = True
        #print("sefl.idea_funnel=", self.idea_funnel, "team_list=", team_list)
        if not self.idea_funnel in team_list:
            print("Wrong Funnel: ", self.idea_funnel)    
            isValid = False
        if not self.idea_status in status_list:
            print("Wrong status: ", self.idea_status, ", should be on of the status of ", status_list)
            isValid = False
        
        return isValid
        
    def teamIndex(self):
        for i, team in enumerate(team_list):
            if team in self.idea_funnel:
                return i

    def statusIndex(self):
        for i, status in enumerate(status_list):
            if status in self.idea_status:
                return i
    
    def toJson(self):
        #return {'Idea Date:':self.idea_date, 'Idea Title':self.idea_title,'Idea Status': self.idea_status, 'Idea Funnel':self.idea_funnel, 'Idea number': self.idea_number  }
        return {'A-Idea Number': self.idea_number, 'E-Idea Title' : self.idea_title,'C-Idea Status': self.idea_status,'D-Idea Date:':self.idea_date,  'B-Idea Funnel':self.idea_funnel }
            
    def printMe(self):
        #print("Idea:%s;%s;%s;%s" % (self.idea_number,self.idea_title,self.idea_funnel, self.idea_status ))
        print("Idea:%s;%s;%s;%s" % (self.idea_number,self.idea_funnel, self.idea_status, self.idea_date ))

#main
if __name__ == "__main__":
    construct_datetimes_range()
    ideasSet=set()
    duplicatedList=list()
    notCalcIdeaList=list()
    #add alll files to data containers
    for file in files:
        addData(file)    
        #verify the data and construct CIdea instance    
        for i in range(len(idea_numbers)):
            idea=CIdea(idea_numbers[i],idea_titles[i], idea_initialtors[i],idea_funnels[i], idea_status[i], 
                   community_discussion[i],
                   implementation_started[i],
                   implementation_done[i]
                  )

            if idea.isInDateRange() and idea.isValid():
                if  ideasSet.add(idea):
                    duplicatedList.append(idea.getNumber())
            else:    
                notCalcIdeaList.append(idea.getNumber())                

    print("Total:", len(ideasSet), "Duplicated ideas:", len(duplicatedList), "Invalid ideas:", len(notCalcIdeaList))
    
    for idea in ideasSet:
        #print(idea.getNumber())
        SCORES[int(idea.statusIndex())][int(idea.teamIndex())] = SCORES[int(idea.statusIndex())][int(idea.teamIndex())] + 1
    
    # calc the total score for each team
    for i in range(len(SCORES[0])):
        SCORES[3][i]=SCORES[1][i]+SCORES[2][i]*3
        
    print(SCORES)
    
    print("type of SCORES:", type(SCORES))
    #saveCompactCsv(compactCsv, ideasSet)
    saveStatisticCsv(statisticCsv, SCORES, ideasSet)
    testChart(compactCsv, SCORES)
    #saveStatisticCsv(compactCsv, SCORES)
    #draw(SCORES)
    #print("scores:", SCORES)
    #scorelist=SCORES.tolist()
    #print("type scores:", type(SCORES))
    #print("type scorelist", type(scorelist))
    #print("scorelist:", scorelist)
    #print("data:", data)
    #print(type(data))
    
    #drawTable(SCORES)
    #drawTable(scorelist)
    #drawTable(data)
